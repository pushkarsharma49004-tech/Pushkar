import streamlit as st
import pandas as pd
import calendar
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Function to generate timesheet DataFrame
def generate_timesheet(emp_name, month, year, leaves, holidays):
    month_days = list(range(1, calendar.monthrange(year, month)[1] + 1))
    weekdays = [calendar.day_name[calendar.weekday(year, month, d)][:3] for d in month_days]

    headers = [f"{d} {w}" for d, w in zip(month_days, weekdays)]
    df = pd.DataFrame(columns=["Emp Name"] + headers)

    row = {"Emp Name": emp_name}
    for d, w in zip(month_days, weekdays):
        if w in ["Sat", "Sun"]:
            row[f"{d} {w}"] = "WO"
        elif d in holidays:
            row[f"{d} {w}"] = "H"
        elif d in leaves:
            row[f"{d} {w}"] = "L"
        else:
            row[f"{d} {w}"] = 1
    df.loc[0] = row
    return df

# Function to style and export Excel
def to_excel_bytes(df, month_name, year, po_number, project_id, tata_manager, client_manager):
    # Step 1: Write DataFrame to BytesIO using pandas
    output = BytesIO()
    df.to_excel(output, engine="openpyxl", index=False, startrow=4)
    output.seek(0)  # Reset the pointer

    # Step 2: Load workbook from BytesIO
    wb = load_workbook(output)
    ws = wb.active

    # Step 3: Modify workbook using openpyxl
    # Header info (PO & Project ID)
    ws["J1"] = f"PO Number: {po_number}"
    ws["J2"] = f"Project ID: {project_id}"
    ws["J3"] = f"Month: {month_name} {year}"
    for cell in ["J1", "J2", "J3"]:
        ws[cell].font = Font(bold=True, color="1F4E78")

    # Style header row
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Borders for all cells
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight weekends, holidays, leaves
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == "WO":
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
            elif cell.value == "H":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
            elif cell.value == "L":
                cell.fill = PatternFill("solid", fgColor="FFD966")
                cell.font = Font(color="7F6000", bold=True)

    # Add signature section
    sign_row = ws.max_row + 3
    ws[f"B{sign_row}"] = "----------------------------------"
    ws[f"B{sign_row+1}"] = " Tata Technologies Manager"
    ws[f"B{sign_row+2}"] = f" Name: {tata_manager}"
    ws[f"B{sign_row+3}"] = " Signature: ___________"
    ws[f"B{sign_row+4}"] = " Date: _______________"
    ws[f"H{sign_row}"] = "----------------------------------"
    ws[f"H{sign_row+1}"] = " Client Manager"
    ws[f"H{sign_row+2}"] = f" Name: {client_manager}"
    ws[f"H{sign_row+3}"] = " Signature: ___________"
    ws[f"H{sign_row+4}"] = " Date: _______________"

    # Step 4: Save workbook to fresh BytesIO
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

# ---------------- Streamlit App ----------------
st.title("📊 Automated Timesheet Generator")

emp_name = st.text_input("Employee Name", "Prabhat Chaurasia")
month = st.number_input("Month (1-12)", 1, 12, 9)
year = st.number_input("Year", 2020, 2100, 2025)
po_number = st.text_input("PO Number", "PO12345")
project_id = st.text_input("Project ID", "PRJ56789")
tata_manager = st.text_input("Tata Technologies Manager", "Shripad")
client_manager = st.text_input("Client Manager", "Client Name")
monthly_rate = st.number_input("Monthly Rate", 0, 1000000, 100000)
contract_days = st.number_input("Contract Days", 1, 31, 22)
leaves = st.text_input("Leave Days (comma separated)", "2, 15")
holidays = st.text_input("Holiday Days (comma separated)", "10, 25")

leave_days = [int(x.strip()) for x in leaves.split(",") if x.strip().isdigit()]
holiday_days = [int(x.strip()) for x in holidays.split(",") if x.strip().isdigit()]
df_timesheet = generate_timesheet(emp_name, month, year, leave_days, holiday_days)

# Billable calculation
worked_days = sum([1 for v in df_timesheet.iloc[0, 1:].values if v == 1])
billable = (worked_days / contract_days) * monthly_rate
st.write(f"### ✅ Worked Days: {worked_days}, Billable Amount: {billable:.2f}")

# Download Excel
month_name = calendar.month_name[month]
if st.button("Generate Timesheet Excel"):
    excel_bytes = to_excel_bytes(df_timesheet, month_name, year, po_number, project_id, tata_manager, client_manager)
    st.download_button("📥 Download Timesheet", excel_bytes, file_name=f"Timesheet_{month_name}_{year}.xlsx")

